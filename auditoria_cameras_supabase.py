import streamlit as st
import pandas as pd
import os
import uuid
import hmac
from datetime import datetime
from io import BytesIO
from urllib.parse import urlparse
import requests
import psycopg2
import psycopg2.extras
from supabase import create_client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as OpenpyxlImage 
from PIL import Image

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as ReportlabImage, PageBreak
    REPORTLAB_OK = True
except Exception:
    REPORTLAB_OK = False

# ── Configuração da Página ──────────────────────────────────────────────────
st.set_page_config(page_title="Auditoria de Câmeras", page_icon="📷", layout="wide")

# ── Configuração de caminhos / Supabase ─────────────────────────────────────
BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
CLIENTES        = os.path.join(BASE_DIR, "nome_clientes.xlsx")
CAMERAS_CSV     = os.path.join(BASE_DIR, "GOV_extracao_cameras.csv")

# No Streamlit Cloud, configure em App > Settings > Secrets:
# SUPABASE_DB_URL = "postgresql://postgres:SUA-SENHA@db.spryvqjlpmmqhtcinslt.supabase.co:5432/postgres"
# SUPABASE_URL = "https://spryvqjlpmmqhtcinslt.supabase.co"
# SUPABASE_SERVICE_ROLE_KEY = "sua-service-role-key"
# SUPABASE_BUCKET = "evidencias"
SUPABASE_DB_URL = st.secrets.get("SUPABASE_DB_URL", os.getenv("SUPABASE_DB_URL", ""))
SUPABASE_URL = st.secrets.get("SUPABASE_URL", os.getenv("SUPABASE_URL", ""))
SUPABASE_KEY = (
    st.secrets.get("SUPABASE_SERVICE_ROLE_KEY", os.getenv("SUPABASE_SERVICE_ROLE_KEY", ""))
    or st.secrets.get("SUPABASE_ANON_KEY", os.getenv("SUPABASE_ANON_KEY", ""))
)
SUPABASE_BUCKET = st.secrets.get("SUPABASE_BUCKET", os.getenv("SUPABASE_BUCKET", "evidencias"))

# ── Login simples do app ────────────────────────────────────────────────────
# Para publicar com segurança, você pode colocar estes valores no Streamlit Secrets:
# APP_USERS = "Fernando,Natanael,Cristina"
# APP_PASSWORD = "sua-senha"
APP_USERS = [
    u.strip().lower()
    for u in str(st.secrets.get("APP_USERS", os.getenv("APP_USERS", "Fernando,Natanael,Cristina"))).split(",")
    if u.strip()
]
APP_PASSWORD = str(st.secrets.get("APP_PASSWORD", os.getenv("APP_PASSWORD", "camerite@123")))


def tela_login():
    """Bloqueia o acesso ao sistema até o usuário autenticar."""
    if st.session_state.get("autenticado"):
        return True

    st.title("🔐 Acesso restrito")
    st.caption("Entre com seu usuário e senha para acessar a Central de Auditoria de Câmeras.")

    with st.form("form_login"):
        usuario = st.text_input("Usuário").strip().lower()
        senha = st.text_input("Senha", type="password")
        entrar = st.form_submit_button("Entrar", type="primary")

    if entrar:
        usuario_ok = usuario in APP_USERS
        senha_ok = hmac.compare_digest(senha, APP_PASSWORD)
        if usuario_ok and senha_ok:
            st.session_state["autenticado"] = True
            st.session_state["usuario_logado"] = usuario
            st.rerun()
        else:
            st.error("Usuário ou senha inválidos.")

    st.stop()


# Fallback local apenas para testes no seu computador.
PASTA_EVIDENCIAS = os.path.join(BASE_DIR, "Evidencias")
PASTA_THUMBNAILS = os.path.join(BASE_DIR, "Evidencias", "Thumbnails")
for pasta in [PASTA_EVIDENCIAS, PASTA_THUMBNAILS]:
    os.makedirs(pasta, exist_ok=True)

def get_db_conn():
    if not SUPABASE_DB_URL:
        st.error("SUPABASE_DB_URL não configurado. Configure em Secrets no Streamlit Cloud.")
        st.stop()
    return psycopg2.connect(SUPABASE_DB_URL, sslmode="require")

@st.cache_resource
def get_supabase_client():
    if not SUPABASE_URL or not SUPABASE_KEY:
        return None
    return create_client(SUPABASE_URL, SUPABASE_KEY)

def caminho_existe(caminho):
    caminho = str(caminho or "").strip()
    if not caminho:
        return False
    if caminho.startswith("http://") or caminho.startswith("https://"):
        return True
    return caminho_existe(caminho)

def abrir_imagem(caminho):
    caminho = str(caminho or "").strip()
    if caminho.startswith("http://") or caminho.startswith("https://"):
        resp = requests.get(caminho, timeout=20)
        resp.raise_for_status()
        return Image.open(BytesIO(resp.content))
    return Image.open(caminho)

# ── Perguntas da auditoria ───────────────────────────────────────────────────
PERGUNTAS = [
    {"chave": "Marca d'Água Travada", "texto": "A marca d'água está travada?", "dica": "⚠️ SIM = marca travada = RUIM", "ruim": "SIM"},
    {"chave": "Câmera está com um bom foco", "texto": "A câmera está com um bom foco?", "dica": "⚠️ NÃO = foco ruim = RUIM", "ruim": "NÃO"},
    {"chave": "Câmera está bem posicionada", "texto": "A câmera está bem posicionada?", "dica": "⚠️ NÃO = posicionamento ruim = RUIM", "ruim": "NÃO"},
    {"chave": "LPR lendo de forma efetiva", "texto": "A LPR está lendo de forma efetiva?", "dica": "ℹ️ Se a câmera NÃO for LPR, responda SIM", "ruim": "NÃO"},
]

# ── Funções de Banco de Dados (Supabase/Postgres) ───────────────────────────
def inicializar_db():
    """Cria as tabelas no Supabase/Postgres se ainda não existirem."""
    conn = get_db_conn()
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS tbl_auditoria (
            id_da_camera TEXT PRIMARY KEY,
            data_auditoria TEXT,
            id_whitelabel TEXT,
            franqueado TEXT,
            cidade TEXT,
            uf TEXT,
            nome_da_camera TEXT,
            status_da_camera TEXT,
            plano_contratado TEXT,
            marca_agua_travada TEXT,
            foco_bom TEXT,
            posicionamento_bom TEXT,
            lpr_efetiva TEXT,
            resultado_geral TEXT,
            observacoes TEXT,
            caminho_evidencia TEXT,
            caminho_thumbnail TEXT
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS tbl_evidencias (
            id BIGSERIAL PRIMARY KEY,
            id_da_camera TEXT NOT NULL,
            caminho_evidencia TEXT NOT NULL,
            caminho_thumbnail TEXT,
            data_upload TEXT,
            observacao TEXT
        )
    """)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_evidencias_camera ON tbl_evidencias (id_da_camera)")
    conn.commit()
    cursor.close()
    conn.close()


def garantir_colunas_dataframe(df, colunas, valor_padrao=""):
    """Garante que o DataFrame tenha todas as colunas esperadas, evitando KeyError."""
    if df is None:
        return pd.DataFrame(columns=colunas)

    for coluna in colunas:
        if coluna not in df.columns:
            df[coluna] = valor_padrao

    return df


def obter_coluna_existente(df, opcoes):
    """Retorna o primeiro nome de coluna encontrado no DataFrame, considerando variações comuns."""
    if df is None or df.empty:
        return None

    mapa = {str(c).strip().lower(): c for c in df.columns}

    for opcao in opcoes:
        chave = str(opcao).strip().lower()
        if chave in mapa:
            return mapa[chave]

    return None


def salvar_ou_atualizar_auditoria(dados):
    conn = get_db_conn()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO tbl_auditoria (
            id_da_camera, data_auditoria, id_whitelabel, franqueado, cidade, uf,
            nome_da_camera, status_da_camera, plano_contratado,
            marca_agua_travada, foco_bom, posicionamento_bom, lpr_efetiva,
            resultado_geral, observacoes, caminho_evidencia, caminho_thumbnail
        ) VALUES (
            %(id_da_camera)s, %(data_auditoria)s, %(id_whitelabel)s, %(franqueado)s, %(cidade)s, %(uf)s,
            %(nome_da_camera)s, %(status_da_camera)s, %(plano_contratado)s,
            %(marca_agua_travada)s, %(foco_bom)s, %(posicionamento_bom)s, %(lpr_efetiva)s,
            %(resultado_geral)s, %(observacoes)s, %(caminho_evidencia)s, %(caminho_thumbnail)s
        )
        ON CONFLICT (id_da_camera) DO UPDATE SET
            data_auditoria = EXCLUDED.data_auditoria,
            id_whitelabel = EXCLUDED.id_whitelabel,
            franqueado = EXCLUDED.franqueado,
            cidade = EXCLUDED.cidade,
            uf = EXCLUDED.uf,
            nome_da_camera = EXCLUDED.nome_da_camera,
            status_da_camera = EXCLUDED.status_da_camera,
            plano_contratado = EXCLUDED.plano_contratado,
            marca_agua_travada = EXCLUDED.marca_agua_travada,
            foco_bom = EXCLUDED.foco_bom,
            posicionamento_bom = EXCLUDED.posicionamento_bom,
            lpr_efetiva = EXCLUDED.lpr_efetiva,
            resultado_geral = EXCLUDED.resultado_geral,
            observacoes = EXCLUDED.observacoes,
            caminho_evidencia = EXCLUDED.caminho_evidencia,
            caminho_thumbnail = EXCLUDED.caminho_thumbnail
    """, {
        "id_da_camera": dados["ID_da_Camera"],
        "data_auditoria": dados["Data_Auditoria"],
        "id_whitelabel": dados["ID_Whitelabel"],
        "franqueado": dados["Franqueado"],
        "cidade": dados["Cidade"],
        "uf": dados["UF"],
        "nome_da_camera": dados["Nome_da_Camera"],
        "status_da_camera": dados["Status_da_Camera"],
        "plano_contratado": dados["Plano_Contratado"],
        "marca_agua_travada": dados["Marca d'Água Travada"],
        "foco_bom": dados["Câmera está com um bom foco"],
        "posicionamento_bom": dados["Câmera está bem posicionada"],
        "lpr_efetiva": dados["LPR lendo de forma efetiva"],
        "resultado_geral": dados["Resultado_Geral"],
        "observacoes": dados["Observacoes"],
        "caminho_evidencia": dados.get("Caminho_Evidencia", ""),
        "caminho_thumbnail": dados.get("Caminho_Thumbnail", ""),
    })
    conn.commit()
    cursor.close()
    conn.close()

def listar_evidencias(id_camera):
    conn = get_db_conn()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cursor.execute("""
        SELECT
            id AS "ID",
            id_da_camera AS "ID_da_Camera",
            caminho_evidencia AS "Caminho_Evidencia",
            caminho_thumbnail AS "Caminho_Thumbnail",
            data_upload AS "Data_Upload",
            observacao AS "Observacao"
        FROM tbl_evidencias
        WHERE id_da_camera = %s
        ORDER BY id ASC
    """, (str(id_camera),))
    registros = [dict(row) for row in cursor.fetchall()]
    cursor.close()
    conn.close()
    return registros

def contar_evidencias(id_camera):
    try:
        conn = get_db_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM tbl_evidencias WHERE id_da_camera = %s", (str(id_camera),))
        total = cursor.fetchone()[0]
        cursor.close()
        conn.close()
        return int(total)
    except Exception:
        return 0

def sincronizar_evidencia_principal(id_camera):
    """Mantém compatibilidade com relatórios antigos usando a primeira evidência como principal."""
    evidencias = listar_evidencias(id_camera)
    caminho = ""
    thumb = ""
    for ev in evidencias:
        caminho_ev = str(ev.get("Caminho_Evidencia", "")).strip()
        if caminho_existe(caminho_ev):
            caminho = caminho_ev
            thumb = str(ev.get("Caminho_Thumbnail", "")).strip()
            break

    conn = get_db_conn()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE tbl_auditoria SET caminho_evidencia = %s, caminho_thumbnail = %s WHERE id_da_camera = %s",
        (caminho, thumb, str(id_camera))
    )
    conn.commit()
    cursor.close()
    conn.close()

def adicionar_evidencia(id_camera, caminho_foto, caminho_thumbnail="", observacao=""):
    conn = get_db_conn()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO tbl_evidencias (id_da_camera, caminho_evidencia, caminho_thumbnail, data_upload, observacao)
        VALUES (%s, %s, %s, %s, %s)
    """, (
        str(id_camera), caminho_foto, caminho_thumbnail,
        datetime.now().strftime("%d/%m/%Y %H:%M"), observacao
    ))
    conn.commit()
    cursor.close()
    conn.close()
    sincronizar_evidencia_principal(id_camera)

def atualizar_apenas_evidencia(id_camera, caminho_foto, caminho_thumbnail=""):
    # Compatibilidade: agora este método acrescenta uma nova evidência em vez de substituir.
    adicionar_evidencia(id_camera, caminho_foto, caminho_thumbnail)


def remover_evidencia_unica(id_evidencia):
    conn = get_db_conn()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cursor.execute("""
        SELECT
            id AS "ID",
            id_da_camera AS "ID_da_Camera",
            caminho_evidencia AS "Caminho_Evidencia",
            caminho_thumbnail AS "Caminho_Thumbnail",
            data_upload AS "Data_Upload",
            observacao AS "Observacao"
        FROM tbl_evidencias
        WHERE id = %s
    """, (id_evidencia,))
    row = cursor.fetchone()
    if not row:
        cursor.close()
        conn.close()
        return None

    dados = dict(row)
    cursor.execute("DELETE FROM tbl_evidencias WHERE id = %s", (id_evidencia,))
    conn.commit()
    cursor.close()
    conn.close()
    sincronizar_evidencia_principal(dados["ID_da_Camera"])
    return dados

def remover_evidencia(id_camera):
    """Remove todas as evidências da câmera selecionada no banco."""
    evidencias = listar_evidencias(id_camera)
    conn = get_db_conn()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_evidencias WHERE id_da_camera = %s", (str(id_camera),))
    cursor.execute(
        "UPDATE tbl_auditoria SET caminho_evidencia = '', caminho_thumbnail = '' WHERE id_da_camera = %s",
        (str(id_camera),)
    )
    conn.commit()
    cursor.close()
    conn.close()
    return evidencias

def migrar_evidencia_antiga_para_galeria(id_camera, caminho_foto, caminho_thumbnail=""):
    """Compatibilidade com base antiga."""
    caminho = str(caminho_foto).strip() if caminho_foto else ""
    if not caminho or not caminho_existe(caminho):
        return

    if contar_evidencias(id_camera) > 0:
        return

    adicionar_evidencia(id_camera, caminho, str(caminho_thumbnail).strip() if caminho_thumbnail else "")

def carregar_imagem_otimizada(caminho, largura_max=1200):
    """Carrega uma imagem em tamanho reduzido para evitar travamentos no Streamlit."""
    try:
        with abrir_imagem(caminho) as img:
            img = img.convert("RGB")
            img.thumbnail((largura_max, largura_max), Image.LANCZOS)
            return img.copy()
    except Exception as e:
        st.error(f"Erro ao carregar imagem: {e}")
        return None

def preparar_imagem_para_excel(caminho, largura_max=900):
    """Cria uma imagem temporária em memória, reduzida, para embutir no Excel sem pesar o app."""
    try:
        with abrir_imagem(caminho) as img:
            img = img.convert("RGB")
            img.thumbnail((largura_max, largura_max), Image.LANCZOS)
            buffer = BytesIO()
            img.save(buffer, format="JPEG", quality=70, optimize=True)
            buffer.seek(0)
            return buffer
    except Exception:
        return None

def salvar_evidencia_otimizada(arquivo_upload, id_camera):
    """Salva a evidência no Supabase Storage e retorna URLs públicas."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    identificador_unico = uuid.uuid4().hex[:8]
    nome_base = f"evidencia_{id_camera}_{timestamp}_{identificador_unico}"
    caminho_storage = f"{id_camera}/{nome_base}.jpg"
    caminho_thumb_storage = f"{id_camera}/{nome_base}_thumb.jpg"

    with Image.open(arquivo_upload) as img:
        img = img.convert("RGB")

        img_principal = img.copy()
        img_principal.thumbnail((1600, 1600), Image.LANCZOS)
        buffer_principal = BytesIO()
        img_principal.save(buffer_principal, "JPEG", quality=78, optimize=True)
        buffer_principal.seek(0)

        img_thumb = img.copy()
        img_thumb.thumbnail((420, 420), Image.LANCZOS)
        buffer_thumb = BytesIO()
        img_thumb.save(buffer_thumb, "JPEG", quality=70, optimize=True)
        buffer_thumb.seek(0)

    supabase = get_supabase_client()
    if supabase is not None:
        supabase.storage.from_(SUPABASE_BUCKET).upload(
            caminho_storage,
            buffer_principal.getvalue(),
            {"content-type": "image/jpeg", "upsert": "false"}
        )
        supabase.storage.from_(SUPABASE_BUCKET).upload(
            caminho_thumb_storage,
            buffer_thumb.getvalue(),
            {"content-type": "image/jpeg", "upsert": "false"}
        )
        url_final = supabase.storage.from_(SUPABASE_BUCKET).get_public_url(caminho_storage)
        url_thumb = supabase.storage.from_(SUPABASE_BUCKET).get_public_url(caminho_thumb_storage)
        return url_final, url_thumb

    # Fallback local para teste fora do Streamlit Cloud.
    caminho_final = os.path.join(PASTA_EVIDENCIAS, f"{nome_base}.jpg")
    caminho_thumb = os.path.join(PASTA_THUMBNAILS, f"{nome_base}_thumb.jpg")
    with open(caminho_final, "wb") as f:
        f.write(buffer_principal.getvalue())
    with open(caminho_thumb, "wb") as f:
        f.write(buffer_thumb.getvalue())
    return caminho_final, caminho_thumb

def excluir_arquivo_se_existir(caminho):
    """Remove arquivo local. Para URL do Supabase, mantém o arquivo no Storage e remove apenas o registro do banco."""
    try:
        caminho_limpo = str(caminho).strip() if caminho else ""
        if caminho_limpo.startswith("http://") or caminho_limpo.startswith("https://"):
            return
        if caminho_limpo and os.path.exists(caminho_limpo):
            os.remove(caminho_limpo)
    except Exception:
        pass

def obter_thumbnail(row_data):
    thumb_raw = row_data.get("Caminho_Thumbnail", "")
    thumb = str(thumb_raw).strip() if pd.notna(thumb_raw) else ""
    if thumb and caminho_existe(thumb):
        return thumb

    img_raw = row_data.get("Caminho_Evidencia", "")
    caminho_img = str(img_raw).strip() if pd.notna(img_raw) else ""
    if caminho_img and caminho_existe(caminho_img):
        if caminho_img.startswith("http://") or caminho_img.startswith("https://"):
            return caminho_img
        try:
            base = os.path.splitext(os.path.basename(caminho_img))[0]
            caminho_thumb = os.path.join(PASTA_THUMBNAILS, f"{base}_thumb.jpg")
            with Image.open(caminho_img) as img:
                img = img.convert("RGB")
                img.thumbnail((420, 420), Image.LANCZOS)
                img.save(caminho_thumb, "JPEG", quality=70, optimize=True)
            return caminho_thumb
        except Exception:
            return caminho_img

    return ""

def carregar_todos_registros():
    conn = get_db_conn()
    df = pd.read_sql_query("""
        SELECT
            data_auditoria AS "Data_Auditoria",
            id_whitelabel AS "ID_Whitelabel",
            franqueado AS "Franqueado",
            cidade AS "Cidade",
            uf AS "UF",
            id_da_camera AS "ID_da_Camera",
            nome_da_camera AS "Nome_da_Camera",
            status_da_camera AS "Status_da_Camera",
            plano_contratado AS "Plano_Contratado",
            marca_agua_travada AS "Marca d'Água Travada",
            foco_bom AS "Câmera está com um bom foco",
            posicionamento_bom AS "Câmera está bem posicionada",
            lpr_efetiva AS "LPR lendo de forma efetiva",
            resultado_geral AS "Resultado_Geral",
            observacoes AS "Observacoes",
            caminho_evidencia AS "Caminho_Evidencia",
            caminho_thumbnail AS "Caminho_Thumbnail"
        FROM tbl_auditoria
    """, conn)
    conn.close()

    df = garantir_colunas_dataframe(df, [
        "Data_Auditoria", "ID_Whitelabel", "Franqueado", "Cidade", "UF",
        "ID_da_Camera", "Nome_da_Camera", "Status_da_Camera", "Plano_Contratado",
        "Marca d'Água Travada", "Câmera está com um bom foco",
        "Câmera está bem posicionada", "LPR lendo de forma efetiva",
        "Resultado_Geral", "Observacoes", "Caminho_Evidencia", "Caminho_Thumbnail"
    ])

    return df

# O banco só é inicializado depois do login.

# ── Carregamento dos dados das Câmeras e Clientes ───────────────────────────
@st.cache_data
def carregar_arquivos_origem():
    try:
        clientes_df = pd.read_excel(CLIENTES, dtype={"ID_Whitelabel": str})
        clientes_df.columns = [str(c).replace("\ufeff", "").strip() for c in clientes_df.columns]

        if "ID_Whitelabel" not in clientes_df.columns:
            return None, "A planilha nome_clientes.xlsx precisa ter a coluna ID_Whitelabel."

        clientes_df["ID_Whitelabel"] = clientes_df["ID_Whitelabel"].astype(str).str.strip()

        # Aceita variações comuns do campo cidade/UF na planilha nome_clientes.xlsx.
        mapa_clientes = {str(c).strip().lower(): c for c in clientes_df.columns}
        col_cidade = (
            mapa_clientes.get("cidade")
            or mapa_clientes.get("cidade/uf")
            or mapa_clientes.get("municipio")
            or mapa_clientes.get("município")
        )
        col_uf = mapa_clientes.get("uf") or mapa_clientes.get("estado")

        clientes_df["cidade"] = clientes_df[col_cidade].fillna("").astype(str) if col_cidade else ""
        clientes_df["uf"] = clientes_df[col_uf].fillna("").astype(str) if col_uf else ""

        meus_ids = set(clientes_df["ID_Whitelabel"].astype(str).str.strip())

        # O CSV padrão GOV enviado é separado por ponto e vírgula (;).
        # Usar sep=None pode falhar ou interpretar errado em alguns ambientes.
        try:
            cameras_df = pd.read_csv(
                CAMERAS_CSV,
                encoding="utf-8-sig",
                sep=";",
                dtype={"ID_Whitelabel": str, "ID_da_Camera": str}
            )
        except Exception:
            cameras_df = pd.read_csv(
                CAMERAS_CSV,
                encoding="latin1",
                sep=";",
                dtype={"ID_Whitelabel": str, "ID_da_Camera": str}
            )

        cameras_df.columns = [str(c).replace("\ufeff", "").strip() for c in cameras_df.columns]

        # Diagnóstico amigável caso o arquivo venha com outro separador.
        if len(cameras_df.columns) == 1 and ";" in str(cameras_df.columns[0]):
            return None, "O CSV não foi separado corretamente. Confirme se o arquivo está delimitado por ponto e vírgula (;)."

        colunas_minimas = ["ID_Whitelabel", "ID_da_Camera", "Nome_da_Camera", "Status_da_Camera"]
        faltantes = [c for c in colunas_minimas if c not in cameras_df.columns]
        if faltantes:
            return None, f"O CSV GOV_extracao_cameras.csv está sem as colunas obrigatórias: {', '.join(faltantes)}"

        if "Plano_Contratado" not in cameras_df.columns:
            cameras_df["Plano_Contratado"] = ""

        if "Nome_Empresa" not in cameras_df.columns:
            cameras_df["Nome_Empresa"] = ""

        cameras_df["ID_Whitelabel"] = cameras_df["ID_Whitelabel"].astype(str).str.strip()
        cameras_df["ID_da_Camera"] = cameras_df["ID_da_Camera"].astype(str).str.strip()
        cameras_df["Nome_da_Camera"] = cameras_df["Nome_da_Camera"].fillna("").astype(str).str.strip()
        cameras_df["Status_da_Camera"] = cameras_df["Status_da_Camera"].fillna("").astype(str).str.upper().str.strip()
        cameras_df["Plano_Contratado"] = cameras_df["Plano_Contratado"].fillna("").astype(str).str.strip()
        cameras_df["Nome_Empresa"] = cameras_df["Nome_Empresa"].fillna("").astype(str).str.strip()

        cameras_df = cameras_df[cameras_df["ID_Whitelabel"].isin(meus_ids)].reset_index(drop=True)

        if cameras_df.empty:
            return None, "Nenhuma câmera encontrada para os clientes da planilha nome_clientes.xlsx."

        clientes_merge = clientes_df[["ID_Whitelabel", "cidade", "uf"]].astype(str).drop_duplicates("ID_Whitelabel")

        cameras_df = cameras_df.merge(
            clientes_merge,
            on="ID_Whitelabel",
            how="left"
        )

        cameras_df["cidade"] = cameras_df["cidade"].fillna("").astype(str).str.strip()
        cameras_df["uf"] = cameras_df["uf"].fillna("").astype(str).str.strip()

        # Regra definitiva:
        # Franqueado/Cliente exibido = Nome_Empresa do CSV.
        # Se Nome_Empresa vier vazio, usa cidade da planilha nome_clientes.xlsx pelo ID_Whitelabel.
        cameras_df["Franqueado"] = cameras_df["Nome_Empresa"].where(
            cameras_df["Nome_Empresa"].astype(str).str.strip() != "",
            cameras_df["cidade"]
        )

        cameras_df["Franqueado"] = cameras_df["Franqueado"].where(
            cameras_df["Franqueado"].astype(str).str.strip() != "",
            "Cliente " + cameras_df["ID_Whitelabel"].astype(str)
        )

        # Blindagem final para evitar KeyError em qualquer tela.
        for coluna in [
            "ID_Whitelabel", "Franqueado", "Nome_Empresa", "ID_da_Camera", "Nome_da_Camera",
            "Status_da_Camera", "Plano_Contratado", "cidade", "uf"
        ]:
            if coluna not in cameras_df.columns:
                cameras_df[coluna] = ""

        return cameras_df, None

    except Exception as e:
        return None, f"Erro ao carregar arquivos: {e}"

# ── Exportar para Excel com Imagens Embutidas ───────────────────────────────
def gerar_excel(df_registros):
    wb = Workbook()
    AZUL_HEADER = "1F4E79"
    VERDE       = "C6EFCE"
    VERMELHO    = "FFC7CE"
    CINZA_GRUPO = "D9EAF7"
    
    ws = wb.active
    ws.title = "Auditoria Geral"

    colunas = ["Data_Auditoria", "ID_Whitelabel", "Franqueado", "Cidade", "UF",
               "ID_da_Camera", "Nome_da_Camera", "Status_da_Camera", "Plano_Contratado",
               "Marca d'Água Travada", "Câmera está com um bom foco", "Câmera está bem posicionada", 
               "LPR lendo de forma efetiva", "Resultado_Geral", "Observacoes"]
    
    df = df_registros.copy()

    for col, nome in enumerate(colunas, 1):
        c = ws.cell(row=1, column=col, value=nome)
        c.font      = Font(bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=AZUL_HEADER)
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_n, row_data in df.reset_index(drop=True).iterrows():
        numero_linha = row_n + 2
        ws.row_dimensions[numero_linha].height = 24

        for col_idx, col_nome in enumerate(colunas, 1):
            val = row_data.get(col_nome, "")
            c = ws.cell(row=numero_linha, column=col_idx, value=val)
            c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")

        col_res = colunas.index("Resultado_Geral") + 1
        cel = ws.cell(row=numero_linha, column=col_res)
        cor = VERDE if cel.value == "APROVADA" else VERMELHO
        cel.fill = PatternFill("solid", fgColor=cor)
        cel.font = Font(bold=True)
        cel.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, _ in enumerate(colunas, 1):
        letra_col = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letra_col].width = 18

    ws_fotos = wb.create_sheet("Relatório Fotográfico")
    ws_fotos.cell(row=1, column=1, value="Evidência Visual (Print)").font = Font(bold=True, color="FFFFFF")
    ws_fotos.cell(row=1, column=1).fill = PatternFill("solid", fgColor=AZUL_HEADER)
    ws_fotos.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    ws_fotos.cell(row=1, column=2, value="Detalhamento da Câmera").font = Font(bold=True, color="FFFFFF")
    ws_fotos.cell(row=1, column=2).fill = PatternFill("solid", fgColor=AZUL_HEADER)
    ws_fotos.cell(row=1, column=2).alignment = Alignment(horizontal="left")
    ws_fotos.column_dimensions["A"].width = 50  
    ws_fotos.column_dimensions["B"].width = 60  
    
    linha_foto_atual = 2
    df_apenas_reprovadas = df[df["Resultado_Geral"] == "REPROVADA"].copy()
    
    for _, row_data in df_apenas_reprovadas.iterrows():
        evidencias = listar_evidencias(row_data.get("ID_da_Camera", ""))
        caminhos_validos = []

        for ev in evidencias:
            caminho_ev = str(ev.get("Caminho_Evidencia", "")).strip()
            if caminho_ev and caminho_existe(caminho_ev):
                caminhos_validos.append((caminho_ev, ev.get("Data_Upload", ""), ev.get("Observacao", "")))

        # Compatibilidade com base antiga que tinha apenas uma imagem por câmera.
        if not caminhos_validos:
            img_raw = row_data.get("Caminho_Evidencia", "")
            caminho_img = str(img_raw).strip() if pd.notna(img_raw) else ""
            if caminho_img and caminho_existe(caminho_img):
                caminhos_validos.append((caminho_img, "", ""))

        if not caminhos_validos:
            caminhos_validos.append(("", "", ""))

        texto_detalhe = (
            f"ID DA CÂMERA: {row_data.get('ID_da_Camera', '')}\n"
            f"NOME DA CÂMERA: {row_data.get('Nome_da_Camera', '')}\n"
            f"CLIENTE / FRANQUEADO: {row_data.get('Franqueado', '')}\n"
            f"DATA DA AUDITORIA: {row_data.get('Data_Auditoria', '')}\n"
            f"TOTAL DE EVIDÊNCIAS: {len([c for c, _, _ in caminhos_validos if c])}\n"
            f"OBSERVAÇÕES: {row_data.get('Observacoes', 'Sem observações.')}"
        )

        for indice_foto, (caminho_img, data_upload, obs_foto) in enumerate(caminhos_validos, start=1):
            ws_fotos.row_dimensions[linha_foto_atual].height = 170
            if caminho_img and caminho_existe(caminho_img):
                try:
                    img_buffer = preparar_imagem_para_excel(caminho_img, largura_max=900)
                    if img_buffer is None:
                        raise ValueError("Imagem inválida")
                    img_relatorio = OpenpyxlImage(img_buffer)
                    img_relatorio.width = 320
                    img_relatorio.height = 180
                    ws_fotos.add_image(img_relatorio, f"A{linha_foto_atual}")
                except Exception:
                    ws_fotos.cell(row=linha_foto_atual, column=1, value="Erro ao renderizar imagem")
            else:
                ws_fotos.cell(row=linha_foto_atual, column=1, value="Sem evidência anexada")

            celula_texto = ws_fotos.cell(row=linha_foto_atual, column=2, value=texto_detalhe if indice_foto == 1 else "")
            celula_texto.font = Font(name="Calibri", size=11, bold=False)
            celula_texto.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
            linha_foto_atual += 1

        linha_foto_atual += 1

    if linha_foto_atual == 2:
        ws_fotos.cell(row=2, column=1, value="Nenhuma evidência fotográfica registrada ou pendente de reprovação.")
        ws_fotos.merge_cells("A2:B2")

    ws2 = wb.create_sheet("Resumo por Cliente")
    ws2.append(["ID_Whitelabel", "Franqueado", "Total Auditadas", "Aprovadas", "Reprovadas", "% Aprovação"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=AZUL_HEADER)
        cell.alignment = Alignment(horizontal="center")

    for (id_wl, franq), grp in df.groupby(["ID_Whitelabel", "Franqueado"]):
        total   = len(grp)
        aprov   = (grp["Resultado_Geral"] == "APROVADA").sum()
        reprov  = total - aprov
        pct     = round(aprov / total * 100, 1) if total else 0
        ws2.append([id_wl, franq, total, aprov, reprov, f"{pct}%"])

    for col in ["A", "B", "C", "D", "E", "F"]:
        ws2.column_dimensions[col].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def gerar_pdf(df_registros):
    if not REPORTLAB_OK:
        raise RuntimeError("Biblioteca reportlab não instalada. Execute: pip install reportlab")

    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TituloAuditoria", parent=styles["Title"], fontSize=18, leading=22, spaceAfter=12)
    normal = styles["Normal"]
    small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=8, leading=10)

    story = []
    story.append(Paragraph("Relatório de Auditoria de Câmeras", title_style))
    story.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", normal))
    story.append(Spacer(1, 0.3 * cm))

    df = df_registros.copy()
    total = len(df)
    aprov = int((df["Resultado_Geral"] == "APROVADA").sum()) if total else 0
    reprov = int((df["Resultado_Geral"] == "REPROVADA").sum()) if total else 0
    pct = round((aprov / total) * 100, 1) if total else 0

    resumo = Table([
        ["Total", "Aprovadas", "Reprovadas", "% Aprovação"],
        [str(total), str(aprov), str(reprov), f"{pct}%"],
    ], colWidths=[4 * cm, 4 * cm, 4 * cm, 4 * cm])
    resumo.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))
    story.append(resumo)
    story.append(Spacer(1, 0.5 * cm))

    story.append(Paragraph("Resumo por Cliente", styles["Heading2"]))
    tabela_clientes = [["ID", "Franqueado", "Total", "Aprov.", "Reprov.", "%"]]
    if not df.empty:
        for (id_wl, franq), grp in df.groupby(["ID_Whitelabel", "Franqueado"], dropna=False):
            t = len(grp)
            a = int((grp["Resultado_Geral"] == "APROVADA").sum())
            r = t - a
            p = round((a / t) * 100, 1) if t else 0
            tabela_clientes.append([str(id_wl), str(franq)[:32], str(t), str(a), str(r), f"{p}%"])

    tab = Table(tabela_clientes, colWidths=[2.2 * cm, 6.0 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm], repeatRows=1)
    tab.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(tab)

    df_reprovadas = df[df["Resultado_Geral"] == "REPROVADA"].copy() if not df.empty else pd.DataFrame()
    if not df_reprovadas.empty:
        story.append(PageBreak())
        story.append(Paragraph("Relatório Fotográfico - Câmeras Reprovadas", styles["Heading2"]))

        for _, row in df_reprovadas.iterrows():
            evidencias = listar_evidencias(row.get("ID_da_Camera", ""))
            caminhos_validos = []
            for ev in evidencias:
                caminho_ev = str(ev.get("Caminho_Evidencia", "")).strip()
                if caminho_ev and caminho_existe(caminho_ev):
                    caminhos_validos.append((caminho_ev, ev.get("Data_Upload", ""), ev.get("Observacao", "")))

            if not caminhos_validos:
                caminho_antigo = str(row.get("Caminho_Evidencia", "")).strip() if pd.notna(row.get("Caminho_Evidencia", "")) else ""
                if caminho_antigo and caminho_existe(caminho_antigo):
                    caminhos_validos.append((caminho_antigo, "", ""))

            if not caminhos_validos:
                caminhos_validos.append(("", "", ""))

            detalhes = (
                f"<b>ID da Câmera:</b> {row.get('ID_da_Camera', '')}<br/>"
                f"<b>Nome:</b> {row.get('Nome_da_Camera', '')}<br/>"
                f"<b>Cliente:</b> {row.get('Franqueado', '')}<br/>"
                f"<b>Cidade/UF:</b> {row.get('Cidade', '')}/{row.get('UF', '')}<br/>"
                f"<b>Data:</b> {row.get('Data_Auditoria', '')}<br/>"
                f"<b>Total de evidências:</b> {len([c for c, _, _ in caminhos_validos if c])}<br/>"
                f"<b>Obs.:</b> {row.get('Observacoes', 'Sem observações.')}"
            )
            story.append(Paragraph(detalhes, small))
            story.append(Spacer(1, 0.15 * cm))

            for caminho, data_upload, obs_foto in caminhos_validos:
                if caminho and caminho_existe(caminho):
                    try:
                        story.append(ReportlabImage(caminho, width=14 * cm, height=7.8 * cm, kind="proportional"))
                    except Exception:
                        story.append(Paragraph("Imagem não pôde ser renderizada.", normal))
                else:
                    story.append(Paragraph("Sem evidência anexada.", normal))
                story.append(Spacer(1, 0.25 * cm))

            story.append(Spacer(1, 0.35 * cm))

    doc.build(story)
    output.seek(0)
    return output


def preparar_df_reprovadas_agrupado(df_registros):
    """Retorna somente reprovadas, ordenadas e prontas para agrupamento por cliente."""
    if df_registros is None or df_registros.empty:
        return pd.DataFrame()

    df = df_registros.copy()
    if "Resultado_Geral" not in df.columns:
        return pd.DataFrame()

    df = df[df["Resultado_Geral"] == "REPROVADA"].copy()
    if df.empty:
        return df

    for col in ["ID_Whitelabel", "Franqueado", "Cidade", "UF", "Nome_da_Camera", "ID_da_Camera"]:
        if col not in df.columns:
            df[col] = ""

    return df.sort_values(["Franqueado", "ID_Whitelabel", "Nome_da_Camera", "ID_da_Camera"]).reset_index(drop=True)


def contar_evidencias_camera(row):
    id_camera = row.get("ID_da_Camera", "")
    evidencias = listar_evidencias(id_camera)
    qtd = 0
    for ev in evidencias:
        caminho = str(ev.get("Caminho_Evidencia", "")).strip()
        if caminho and caminho_existe(caminho):
            qtd += 1

    if qtd == 0:
        caminho_antigo = str(row.get("Caminho_Evidencia", "")).strip() if pd.notna(row.get("Caminho_Evidencia", "")) else ""
        if caminho_antigo and caminho_existe(caminho_antigo):
            qtd = 1
    return qtd


def gerar_excel_reprovadas_agrupadas(df_registros):
    """Gera Excel somente com reprovadas, agrupadas por cliente."""
    wb = Workbook()
    AZUL_HEADER = "1F4E79"
    VERMELHO = "FFC7CE"
    CINZA_GRUPO = "D9EAF7"

    df = preparar_df_reprovadas_agrupado(df_registros)

    ws = wb.active
    ws.title = "Reprovadas por Cliente"
    colunas = [
        "ID_Whitelabel", "Franqueado", "Cidade", "UF", "ID_da_Camera", "Nome_da_Camera",
        "Status_da_Camera", "Plano_Contratado", "Observacoes", "Qtd_Evidencias"
    ]
    ws.append(colunas)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=AZUL_HEADER)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    linha = 2
    if not df.empty:
        for (id_wl, franq), grp in df.groupby(["ID_Whitelabel", "Franqueado"], dropna=False):
            ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=len(colunas))
            c = ws.cell(row=linha, column=1, value=f"CLIENTE: {id_wl} - {franq} | REPROVADAS: {len(grp)}")
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor=CINZA_GRUPO)
            linha += 1

            for _, row in grp.iterrows():
                qtd_evid = contar_evidencias_camera(row.to_dict())
                valores = [
                    row.get("ID_Whitelabel", ""), row.get("Franqueado", ""), row.get("Cidade", ""), row.get("UF", ""),
                    row.get("ID_da_Camera", ""), row.get("Nome_da_Camera", ""), row.get("Status_da_Camera", ""),
                    row.get("Plano_Contratado", ""), row.get("Observacoes", ""), qtd_evid
                ]
                ws.append(valores)
                for col_idx in range(1, len(colunas) + 1):
                    cel = ws.cell(row=linha, column=col_idx)
                    cel.alignment = Alignment(wrap_text=True, vertical="center")
                    if colunas[col_idx - 1] == "ID_da_Camera":
                        cel.fill = PatternFill("solid", fgColor=VERMELHO)
                linha += 1

    for idx in range(1, len(colunas) + 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = 22

    ws_resumo = wb.create_sheet("Resumo")
    ws_resumo.append(["ID_Whitelabel", "Franqueado", "Cidade/UF", "Qtd Reprovadas", "Qtd Evidências"])
    for cell in ws_resumo[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=AZUL_HEADER)
        cell.alignment = Alignment(horizontal="center")

    if not df.empty:
        for (id_wl, franq), grp in df.groupby(["ID_Whitelabel", "Franqueado"], dropna=False):
            cidade_uf = f"{grp.iloc[0].get('Cidade', '')}/{grp.iloc[0].get('UF', '')}"
            total_evid = sum(contar_evidencias_camera(row.to_dict()) for _, row in grp.iterrows())
            ws_resumo.append([id_wl, franq, cidade_uf, len(grp), total_evid])

    for col in ["A", "B", "C", "D", "E"]:
        ws_resumo.column_dimensions[col].width = 24

    ws_fotos = wb.create_sheet("Fotos Reprovadas")
    ws_fotos.append(["Evidência Visual", "Detalhamento"])
    for cell in ws_fotos[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=AZUL_HEADER)
    ws_fotos.column_dimensions["A"].width = 50
    ws_fotos.column_dimensions["B"].width = 70

    linha_foto = 2
    if not df.empty:
        for (id_wl, franq), grp in df.groupby(["ID_Whitelabel", "Franqueado"], dropna=False):
            ws_fotos.merge_cells(start_row=linha_foto, start_column=1, end_row=linha_foto, end_column=2)
            c = ws_fotos.cell(row=linha_foto, column=1, value=f"CLIENTE: {id_wl} - {franq}")
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor=CINZA_GRUPO)
            linha_foto += 1

            for _, row_data in grp.iterrows():
                evidencias = listar_evidencias(row_data.get("ID_da_Camera", ""))
                caminhos_validos = []
                for ev in evidencias:
                    caminho_ev = str(ev.get("Caminho_Evidencia", "")).strip()
                    if caminho_ev and caminho_existe(caminho_ev):
                        caminhos_validos.append((caminho_ev, ev.get("Data_Upload", ""), ev.get("Observacao", "")))

                if not caminhos_validos:
                    caminho_antigo = str(row_data.get("Caminho_Evidencia", "")).strip() if pd.notna(row_data.get("Caminho_Evidencia", "")) else ""
                    if caminho_antigo and caminho_existe(caminho_antigo):
                        caminhos_validos.append((caminho_antigo, "", ""))

                if not caminhos_validos:
                    caminhos_validos.append(("", "", ""))

                detalhe = (
                    f"CLIENTE: {id_wl} - {franq}\n"
                    f"ID DA CÂMERA: {row_data.get('ID_da_Camera', '')}\n"
                    f"NOME DA CÂMERA: {row_data.get('Nome_da_Camera', '')}\n"
                    f"CIDADE/UF: {row_data.get('Cidade', '')}/{row_data.get('UF', '')}\n"
                    f"TOTAL DE EVIDÊNCIAS: {len([c for c, _, _ in caminhos_validos if c])}\n"
                    f"OBSERVAÇÕES: {row_data.get('Observacoes', '')}"
                )

                for idx_foto, (caminho_img, data_upload, obs_foto) in enumerate(caminhos_validos, start=1):
                    ws_fotos.row_dimensions[linha_foto].height = 170
                    if caminho_img and caminho_existe(caminho_img):
                        try:
                            img_buffer = preparar_imagem_para_excel(caminho_img, largura_max=900)
                            if img_buffer:
                                img_relatorio = OpenpyxlImage(img_buffer)
                                img_relatorio.width = 320
                                img_relatorio.height = 180
                                ws_fotos.add_image(img_relatorio, f"A{linha_foto}")
                        except Exception:
                            ws_fotos.cell(row=linha_foto, column=1, value="Erro ao renderizar imagem")
                    else:
                        ws_fotos.cell(row=linha_foto, column=1, value="Sem evidência anexada")

                    cel = ws_fotos.cell(row=linha_foto, column=2, value=detalhe if idx_foto == 1 else "")
                    cel.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
                    linha_foto += 1

                linha_foto += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def gerar_pdf_reprovadas_agrupadas(df_registros):
    """Gera PDF somente com reprovadas, agrupadas por cliente."""
    if not REPORTLAB_OK:
        raise RuntimeError("Biblioteca reportlab não instalada. Execute: pip install reportlab")

    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TituloReprovadas", parent=styles["Title"], fontSize=18, leading=22, spaceAfter=12)
    small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=8, leading=10)
    normal = styles["Normal"]

    df = preparar_df_reprovadas_agrupado(df_registros)
    story = []
    story.append(Paragraph("Relatório de Câmeras Reprovadas por Cliente", title_style))
    story.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", normal))
    story.append(Spacer(1, 0.3 * cm))

    if df.empty:
        story.append(Paragraph("Nenhuma câmera reprovada encontrada.", normal))
        doc.build(story)
        output.seek(0)
        return output

    resumo = [["ID", "Cliente", "Reprovadas", "Evidências"]]
    for (id_wl, franq), grp in df.groupby(["ID_Whitelabel", "Franqueado"], dropna=False):
        total_evid = sum(contar_evidencias_camera(row.to_dict()) for _, row in grp.iterrows())
        resumo.append([str(id_wl), str(franq)[:34], str(len(grp)), str(total_evid)])

    tab_resumo = Table(resumo, colWidths=[2.5 * cm, 8 * cm, 3 * cm, 3 * cm], repeatRows=1)
    tab_resumo.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(tab_resumo)

    for idx_cliente, ((id_wl, franq), grp) in enumerate(df.groupby(["ID_Whitelabel", "Franqueado"], dropna=False), start=1):
        story.append(PageBreak())
        story.append(Paragraph(f"Cliente: {id_wl} - {franq}", styles["Heading2"]))
        story.append(Paragraph(f"Total de câmeras reprovadas: {len(grp)}", normal))
        story.append(Spacer(1, 0.2 * cm))

        tabela_cams = [["ID Câmera", "Nome", "Cidade/UF", "Evid.", "Observação"]]
        for _, row in grp.iterrows():
            tabela_cams.append([
                str(row.get("ID_da_Camera", "")),
                str(row.get("Nome_da_Camera", ""))[:28],
                f"{row.get('Cidade', '')}/{row.get('UF', '')}",
                str(contar_evidencias_camera(row.to_dict())),
                str(row.get("Observacoes", ""))[:38],
            ])
        tab_cams = Table(tabela_cams, colWidths=[2.7 * cm, 5.2 * cm, 2.8 * cm, 1.4 * cm, 5.0 * cm], repeatRows=1)
        tab_cams.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
        ]))
        story.append(tab_cams)
        story.append(Spacer(1, 0.4 * cm))

        for _, row in grp.iterrows():
            evidencias = listar_evidencias(row.get("ID_da_Camera", ""))
            caminhos_validos = []
            for ev in evidencias:
                caminho_ev = str(ev.get("Caminho_Evidencia", "")).strip()
                if caminho_ev and caminho_existe(caminho_ev):
                    caminhos_validos.append((caminho_ev, ev.get("Data_Upload", ""), ev.get("Observacao", "")))

            if not caminhos_validos:
                caminho_antigo = str(row.get("Caminho_Evidencia", "")).strip() if pd.notna(row.get("Caminho_Evidencia", "")) else ""
                if caminho_antigo and caminho_existe(caminho_antigo):
                    caminhos_validos.append((caminho_antigo, "", ""))

            if not caminhos_validos:
                caminhos_validos.append(("", "", ""))

            detalhes = (
                f"<b>ID da Câmera:</b> {row.get('ID_da_Camera', '')}<br/>"
                f"<b>Nome:</b> {row.get('Nome_da_Camera', '')}<br/>"
                f"<b>Cliente:</b> {row.get('Franqueado', '')}<br/>"
                f"<b>Cidade/UF:</b> {row.get('Cidade', '')}/{row.get('UF', '')}<br/>"
                f"<b>Total de evidências:</b> {len([c for c, _, _ in caminhos_validos if c])}<br/>"
                f"<b>Obs. auditoria:</b> {row.get('Observacoes', 'Sem observações.')}"
            )
            story.append(Paragraph(detalhes, small))
            story.append(Spacer(1, 0.15 * cm))

            for caminho, data_upload, obs_foto in caminhos_validos:
                if caminho and caminho_existe(caminho):
                    try:
                        story.append(ReportlabImage(caminho, width=14 * cm, height=7.8 * cm, kind="proportional"))
                    except Exception:
                        story.append(Paragraph("Imagem não pôde ser renderizada.", normal))
                else:
                    story.append(Paragraph("Sem evidência anexada.", normal))
                story.append(Spacer(1, 0.25 * cm))

            story.append(Spacer(1, 0.35 * cm))

    doc.build(story)
    output.seek(0)
    return output

@st.cache_data(ttl=30)
def carregar_reprovadas(df):
    return df[df["Resultado_Geral"] == "REPROVADA"].copy()

def exibir_historico_auditoria_agrupado(df_salvos, id_cliente_selecionado=None):
    """Exibe o histórico somente na aba Realizar Auditoria, com reprovadas primeiro e aprovadas abaixo."""
    st.divider()
    st.subheader("📋 Histórico de Auditoria")

    if df_salvos.empty:
        st.info("Nenhum registro encontrado no banco de dados.")
        return

    df_hist = df_salvos.copy()

    if id_cliente_selecionado:
        id_filtro = str(id_cliente_selecionado).strip()
        df_hist = df_hist[df_hist["ID_Whitelabel"].astype(str).str.strip() == id_filtro]
        st.caption(f"Filtro ativo: cliente `{id_filtro}`")
    else:
        st.caption("Nenhum cliente selecionado: exibindo todos os clientes.")

    if df_hist.empty:
        st.warning("Nenhum histórico encontrado para o cliente selecionado.")
        return

    colunas_ocultar = [
        "Caminho_Evidencia",
        "Caminho_Thumbnail",
    ]

    colunas_preferidas = [
        "Data_Auditoria",
        "ID_Whitelabel",
        "Franqueado",
        "Cidade",
        "UF",
        "ID_da_Camera",
        "Nome_da_Camera",
        "Status_da_Camera",
        "Plano_Contratado",
        "Marca d'Água Travada",
        "Câmera está com um bom foco",
        "Câmera está bem posicionada",
        "LPR lendo de forma efetiva",
        "Resultado_Geral",
        "Observacoes",
    ]

    colunas_existentes = [c for c in colunas_preferidas if c in df_hist.columns]
    outras_colunas = [c for c in df_hist.columns if c not in colunas_existentes and c not in colunas_ocultar]
    df_hist = df_hist[colunas_existentes + outras_colunas]

    df_reprovadas_hist = df_hist[df_hist["Resultado_Geral"] == "REPROVADA"].copy()
    df_aprovadas_hist = df_hist[df_hist["Resultado_Geral"] == "APROVADA"].copy()

    with st.expander(f"❌ Reprovadas ({len(df_reprovadas_hist)})", expanded=True):
        if df_reprovadas_hist.empty:
            st.success("Nenhuma câmera reprovada nesse filtro.")
        else:
            st.dataframe(df_reprovadas_hist, width=1600, hide_index=True)

    with st.expander(f"✅ Aprovadas ({len(df_aprovadas_hist)})", expanded=False):
        if df_aprovadas_hist.empty:
            st.info("Nenhuma câmera aprovada nesse filtro.")
        else:
            st.dataframe(df_aprovadas_hist, width=1600, hide_index=True)



def montar_html_email_reprovadas(df_registros, titulo_filtro="Todos os clientes"):
    """Monta o corpo HTML do e-mail com o conteúdo textual/fotográfico das reprovadas."""
    df = preparar_df_reprovadas_agrupado(df_registros)
    data_geracao = datetime.now().strftime("%d/%m/%Y %H:%M")

    html = []
    html.append("""
    <html>
    <body style="font-family: Calibri, Arial, sans-serif; font-size: 11pt; color: #1f1f1f;">
    """)
    html.append("<p>Prezados,</p>")
    html.append("<p>Segue abaixo o relatório das câmeras reprovadas na auditoria.</p>")
    html.append(f"<p><b>Filtro:</b> {titulo_filtro}<br>")
    html.append(f"<b>Gerado em:</b> {data_geracao}<br>")
    html.append(f"<b>Total de câmeras reprovadas:</b> {len(df)}</p>")

    if df.empty:
        html.append("<p>Nenhuma câmera reprovada encontrada para o filtro selecionado.</p>")
        html.append("</body></html>")
        return "".join(html), []

    html.append("""
    <table style="border-collapse: collapse; width: 100%; margin-bottom: 18px;">
        <tr style="background-color: #1F4E79; color: white;">
            <th style="border: 1px solid #999; padding: 6px; text-align:left;">ID</th>
            <th style="border: 1px solid #999; padding: 6px; text-align:left;">Cliente</th>
            <th style="border: 1px solid #999; padding: 6px; text-align:center;">Reprovadas</th>
            <th style="border: 1px solid #999; padding: 6px; text-align:center;">Evidências</th>
        </tr>
    """)

    for (id_wl, franq), grp in df.groupby(["ID_Whitelabel", "Franqueado"], dropna=False):
        total_evid = sum(contar_evidencias_camera(row.to_dict()) for _, row in grp.iterrows())
        html.append(f"""
        <tr>
            <td style="border: 1px solid #999; padding: 6px;">{id_wl}</td>
            <td style="border: 1px solid #999; padding: 6px;">{franq}</td>
            <td style="border: 1px solid #999; padding: 6px; text-align:center;">{len(grp)}</td>
            <td style="border: 1px solid #999; padding: 6px; text-align:center;">{total_evid}</td>
        </tr>
        """)
    html.append("</table>")

    anexos_inline = []
    idx_img = 1

    for (id_wl, franq), grp in df.groupby(["ID_Whitelabel", "Franqueado"], dropna=False):
        html.append(f"<h2 style='color:#1F4E79; border-bottom:1px solid #ccc;'>Cliente: {id_wl} - {franq}</h2>")
        html.append(f"<p><b>Total de câmeras reprovadas:</b> {len(grp)}</p>")

        for _, row in grp.iterrows():
            id_cam = str(row.get("ID_da_Camera", ""))
            nome_cam = str(row.get("Nome_da_Camera", ""))
            cidade_uf = f"{row.get('Cidade', '')}/{row.get('UF', '')}"
            obs = str(row.get("Observacoes", "") or "")
            evidencias = listar_evidencias(id_cam)

            html.append("""
            <div style="border: 1px solid #d9d9d9; border-radius: 6px; padding: 10px; margin: 12px 0;">
            """)
            html.append(f"<p style='margin:0 0 8px 0;'><b>ID da Câmera:</b> {id_cam}<br>")
            html.append(f"<b>Nome:</b> {nome_cam}<br>")
            html.append(f"<b>Cliente:</b> {franq}<br>")
            html.append(f"<b>Cidade/UF:</b> {cidade_uf}<br>")
            html.append(f"<b>Quantidade de evidências:</b> {len(evidencias)}<br>")
            html.append(f"<b>Obs. auditoria:</b> {obs}</p>")

            caminhos_validos = []
            for ev in evidencias:
                caminho_ev = str(ev.get("Caminho_Evidencia", "")).strip()
                if caminho_ev and caminho_existe(caminho_ev):
                    caminhos_validos.append(caminho_ev)

            if caminhos_validos:
                for caminho_ev in caminhos_validos:
                    cid = f"evidencia_{id_cam}_{idx_img}"
                    anexos_inline.append({"path": caminho_ev, "cid": cid})
                    html.append(f"<div style='margin-top:8px;'><img src='cid:{cid}' style='max-width: 760px; width: 100%; height: auto; border: 1px solid #ccc;'></div>")
                    idx_img += 1
            else:
                html.append("<p><i>Sem evidência fotográfica vinculada.</i></p>")

            html.append("</div>")

    html.append("<p>Atenciosamente,</p>")
    html.append("</body></html>")
    return "".join(html), anexos_inline


def abrir_email_outlook_classico_reprovadas(df_registros, titulo_filtro="Todos os clientes"):
    """Abre um novo e-mail no Outlook clássico com o relatório no corpo, sem assunto automático."""
    if os.name != "nt":
        raise RuntimeError("A abertura direta no Outlook clássico funciona somente no Windows com Outlook instalado.")

    try:
        import win32com.client
    except Exception:
        raise RuntimeError("Biblioteca pywin32 não encontrada. Instale com: pip install pywin32")

    html_body, anexos_inline = montar_html_email_reprovadas(df_registros, titulo_filtro)

    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)
    mail.Subject = ""
    mail.HTMLBody = html_body

    # Anexa as imagens como inline para aparecerem no corpo do e-mail.
    # Propriedade MAPI: PR_ATTACH_CONTENT_ID
    for item in anexos_inline:
        caminho = item["path"]
        cid = item["cid"]
        try:
            anexo = mail.Attachments.Add(Source=caminho)
            anexo.PropertyAccessor.SetProperty(
                "http://schemas.microsoft.com/mapi/proptag/0x3712001F",
                cid
            )
        except Exception:
            # Se algum anexo falhar, o e-mail ainda abre com o conteúdo textual.
            pass

    mail.Display(False)


def exibir_pdf_reprovadas_auditoria(df_salvos, id_cliente_selecionado=None):
    """Gera PDF das reprovadas na aba Realizar Auditoria, apenas sob demanda."""
    st.divider()
    st.subheader("🧾 PDF das Reprovadas")

    if df_salvos.empty:
        st.info("Nenhum registro encontrado para gerar PDF.")
        return

    df_pdf = df_salvos.copy()
    sufixo_arquivo = "geral"
    titulo_filtro = "Todos os clientes"

    if id_cliente_selecionado:
        id_filtro = str(id_cliente_selecionado).strip()
        df_pdf = df_pdf[df_pdf["ID_Whitelabel"].astype(str).str.strip() == id_filtro]
        sufixo_arquivo = f"cliente_{id_filtro}"
        titulo_filtro = f"Cliente {id_filtro}"

    df_reprovadas_pdf = df_pdf[df_pdf["Resultado_Geral"] == "REPROVADA"].copy()

    if df_reprovadas_pdf.empty:
        st.success(f"Nenhuma câmera reprovada encontrada para o filtro: {titulo_filtro}.")
        return

    clientes_envolvidos = df_reprovadas_pdf["ID_Whitelabel"].astype(str).nunique() if "ID_Whitelabel" in df_reprovadas_pdf.columns else 0
    total_reprovadas = len(df_reprovadas_pdf)

    col_pdf_1, col_pdf_2, col_pdf_3 = st.columns(3)
    col_pdf_1.metric("Filtro", titulo_filtro)
    col_pdf_2.metric("Câmeras reprovadas", total_reprovadas)
    col_pdf_3.metric("Clientes no PDF", clientes_envolvidos)

    st.caption("O PDF só é montado quando você clicar no botão abaixo. Isso evita travamento ao navegar entre câmeras.")

    chave_pdf_auditoria = f"pdf_auditoria_{sufixo_arquivo}_{total_reprovadas}_{clientes_envolvidos}"

    if st.session_state.get("chave_pdf_auditoria") != chave_pdf_auditoria:
        st.session_state.pop("pdf_auditoria_data", None)
        st.session_state.pop("pdf_auditoria_nome", None)
        st.session_state["chave_pdf_auditoria"] = chave_pdf_auditoria

    if st.button("🧾 Preparar PDF das Reprovadas deste filtro", type="secondary", key=f"btn_pdf_auditoria_{sufixo_arquivo}"):
        try:
            with st.spinner("Montando PDF das reprovadas..."):
                pdf_data = gerar_pdf_reprovadas_agrupadas(df_pdf)
                nome_pdf = f"reprovadas_{sufixo_arquivo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                st.session_state["pdf_auditoria_data"] = pdf_data
                st.session_state["pdf_auditoria_nome"] = nome_pdf
        except Exception as e:
            st.error(f"Erro ao gerar PDF: {e}")

    if "pdf_auditoria_data" in st.session_state:
        st.download_button(
            label="⬇️ Baixar PDF das Reprovadas deste filtro",
            data=st.session_state["pdf_auditoria_data"],
            file_name=st.session_state.get("pdf_auditoria_nome", "reprovadas.pdf"),
            mime="application/pdf",
            key=f"download_pdf_auditoria_{sufixo_arquivo}"
        )

        st.caption("O botão abaixo abre um novo e-mail no Outlook clássico com o conteúdo do relatório no corpo. O assunto fica em branco para você preencher.")
        if st.button("📧 Abrir e-mail no Outlook com este conteúdo", type="secondary", key=f"btn_email_outlook_{sufixo_arquivo}"):
            try:
                with st.spinner("Abrindo e-mail no Outlook..."):
                    abrir_email_outlook_classico_reprovadas(df_pdf, titulo_filtro)
                st.success("E-mail aberto no Outlook. Revise o conteúdo antes de enviar.")
            except Exception as e:
                st.error(f"Não foi possível abrir o Outlook: {e}")


# ── Interface Principal ──────────────────────────────────────────────────────
def main():
    tela_login()
    inicializar_db()

    st.title("📷 Central de Auditoria de Câmeras")

    cameras_df, erro = carregar_arquivos_origem()
    if erro:
        st.error(erro)
        return

    df_salvos = carregar_todos_registros()

    mapa_resultados_cmeras = {}
    if not df_salvos.empty:
        mapa_resultados_cmeras = dict(zip(df_salvos["ID_da_Camera"].astype(str), df_salvos["Resultado_Geral"]))

    # ── FIX 1: Guardar o cliente selecionado no session_state ──────────────
    # Isso garante que a variável persista quando o usuário troca de aba,
    # evitando o loop de re-render infinito que causava o "carregando eternamente".
    if "id_cliente_selecionado" not in st.session_state:
        st.session_state["id_cliente_selecionado"] = None

    tab_auditoria, tab_evidencias, tab_agrupamento = st.tabs(["💻 Realizar Auditoria", "📸 Anexar Evidências (Apenas Reprovadas)", "📊 Reprovadas por Cliente"])

    # =========================================================================
    # ABA 1: REALIZAR AUDITORIA
    # =========================================================================
    with tab_auditoria:
        clientes_lista = cameras_df[["ID_Whitelabel", "Franqueado"]].drop_duplicates().sort_values("Franqueado")
        lista_clientes_mapeados = []
        
        for _, row in clientes_lista.iterrows():
            id_wl = str(row["ID_Whitelabel"])
            nome_franq = row["Franqueado"]
            total_cameras_origem = cameras_df[cameras_df["ID_Whitelabel"] == id_wl].shape[0]
            total_cameras_salvas = df_salvos[df_salvos["ID_Whitelabel"] == id_wl].shape[0] if not df_salvos.empty else 0
            
            if total_cameras_salvas == 0:
                emoji = "⚪"
            elif total_cameras_salvas >= total_cameras_origem:
                emoji = "🟢"
            else:
                emoji = "🟡"
                
            texto_exibicao = f"{emoji} {id_wl} - {nome_franq}"
            lista_clientes_mapeados.append({"id_wl": id_wl, "texto_exibicao": texto_exibicao, "franqueado": nome_franq})

        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**1️⃣ Filtrar Cliente (Digite ID ou Nome)**")
            busca_cliente = st.text_input("Buscar Cliente:", key="busca_cliente_tab1", placeholder="Ex: Joinville ou 1042...")
            
            if busca_cliente:
                clientes_filtrados = [c for c in lista_clientes_mapeados if busca_cliente.lower() in c["texto_exibicao"].lower()]
            else:
                clientes_filtrados = lista_clientes_mapeados

            if len(clientes_filtrados) == 0:
                st.error("❌ Nenhum cliente encontrado.")
                cliente_selecionado_txt = ""
            else:
                opcoes_combo_cliente = [c["texto_exibicao"] for c in clientes_filtrados]
                index_cliente = 0 if len(opcoes_combo_cliente) == 1 else None
                cliente_selecionado_txt = st.selectbox(
                    "Resultados do filtro de cliente:",
                    options=[""] + opcoes_combo_cliente if index_cliente is None else opcoes_combo_cliente,
                    key="combo_cliente_tab1"
                )

        cams_cliente = pd.DataFrame()
        camera_selecionada_txt = ""

        if cliente_selecionado_txt:
            id_cliente_atual = cliente_selecionado_txt.split(" ")[1]
            # ── FIX 1 (continuação): Atualizar session_state ao selecionar cliente ──
            st.session_state["id_cliente_selecionado"] = id_cliente_atual
            cams_cliente = cameras_df[cameras_df["ID_Whitelabel"] == id_cliente_atual]
            
            lista_cameras_mapeadas = []
            for _, row_cam in cams_cliente.iterrows():
                id_cam = str(row_cam["ID_da_Camera"])
                nome_cam = row_cam["Nome_da_Camera"]
                status_conexao = str(row_cam["Status_da_Camera"]).strip().upper()
                res_auditoria = mapa_resultados_cmeras.get(id_cam, None)
                
                if res_auditoria == "APROVADA":
                    emoji_cam = "✅"
                    status_txt = "Aprovada (Editar)"
                elif res_auditoria == "REPROVADA":
                    emoji_cam = "❌"
                    status_txt = "Reprovada (Editar)"
                else:
                    emoji_cam = "⏳"
                    status_txt = "Pendente"
                    
                texto_cam_exibicao = f"{emoji_cam} [{status_conexao}] {id_cam} - {nome_cam} ({status_txt})"
                lista_cameras_mapeadas.append({"id_camera": id_cam, "texto_exibicao": texto_cam_exibicao})
            
            with col2:
                st.markdown("**2️⃣ Filtrar Câmera (Digite ID ou Nome)**")
                busca_camera = st.text_input("Buscar Câmera:", key="busca_camera_tab1", placeholder="Ex: OFFLINE, ID ou trecho...")
                
                if busca_camera:
                    cameras_filtradas = [c for c in lista_cameras_mapeadas if busca_camera.lower() in c["texto_exibicao"].lower()]
                else:
                    cameras_filtradas = lista_cameras_mapeadas

                if len(cameras_filtradas) == 0:
                    st.error("❌ Nenhuma câmera encontrada.")
                else:
                    opcoes_combo_camera = [c["texto_exibicao"] for c in cameras_filtradas]
                    index_camera = 0 if len(opcoes_combo_camera) == 1 else None
                    camera_selecionada_txt = st.selectbox(
                        "Resultados do filtro de câmera:",
                        options=[""] + opcoes_combo_camera if index_camera is None else opcoes_combo_camera,
                        key="combo_camera_tab1"
                    )

        if cliente_selecionado_txt and camera_selecionada_txt:
            partes_selecao = camera_selecionada_txt.split(" ")
            id_camera = partes_selecao[2]
            camera_row = cams_cliente[cams_cliente["ID_da_Camera"].astype(str) == id_camera].iloc[0]
            
            registro_existente = None
            if not df_salvos.empty and id_camera in df_salvos["ID_da_Camera"].values:
                row_dict = df_salvos[df_salvos["ID_da_Camera"] == id_camera].iloc[0].to_dict()
                registro_existente = {k: ("" if pd.isna(v) else v) for k, v in row_dict.items()}

            st.divider()
            st.subheader(f"Formulário de Auditoria: {camera_row['Nome_da_Camera']}")
            
            col_info1, col_info2 = st.columns(2)
            if camera_row['Status_da_Camera'] == 'OFFLINE':
                col_info1.error(f"🚨 **CONEXÃO:** CÂMERA OFFLINE | **ID:** {camera_row['ID_da_Camera']}")
            else:
                col_info1.success(f"🟢 **CONEXÃO:** CÂMERA ONLINE | **ID:** {camera_row['ID_da_Camera']}")
            col_info2.info(f"**Cidade/UF:** {camera_row.get('cidade','')} / {camera_row.get('uf','')}")

            with st.form(key=f"form_auditoria_{id_camera}"):
                respostas = {}
                for p in PERGUNTAS:
                    st.markdown(f"**{p['texto']}**")
                    valor_padrao = registro_existente.get(p["chave"], "SIM") if registro_existente else "SIM"
                    if valor_padrao not in ["SIM", "NÃO"]:
                        valor_padrao = "SIM"
                    respostas[p["chave"]] = st.radio(
                        "Resposta:", options=["SIM", "NÃO"],
                        index=0 if valor_padrao == "SIM" else 1,
                        key=f"aud_{id_camera}_{p['chave']}", horizontal=True
                    )
                
                obs_padrao = registro_existente.get("Observacoes", "") if registro_existente else ""
                obs = st.text_area("Observações adicionais", value=obs_padrao)
                
                if st.form_submit_button("Salvar Auditoria", type="primary"):
                    reprovada = any(respostas[p["chave"]] == p["ruim"] for p in PERGUNTAS)
                    img_preservada = registro_existente.get("Caminho_Evidencia", "") if registro_existente else ""
                    thumb_preservada = registro_existente.get("Caminho_Thumbnail", "") if registro_existente else ""
                    
                    registro = {
                        "Data_Auditoria": datetime.now().strftime("%d/%m/%Y %H:%M"),
                        "ID_Whitelabel": camera_row["ID_Whitelabel"],
                        "Franqueado": camera_row["Franqueado"],
                        "Cidade": camera_row.get("cidade", ""),
                        "UF": camera_row.get("uf", ""),
                        "ID_da_Camera": camera_row["ID_da_Camera"],
                        "Nome_da_Camera": camera_row["Nome_da_Camera"],
                        "Status_da_Camera": camera_row["Status_da_Camera"],
                        "Plano_Contratado": camera_row.get("Plano_Contratado", ""),
                        "Caminho_Evidencia": img_preservada if reprovada else "",
                        "Caminho_Thumbnail": thumb_preservada if reprovada else "",
                        "Observacoes": obs,
                        "Resultado_Geral": "REPROVADA" if reprovada else "APROVADA"
                    }
                    registro.update(respostas)
                    salvar_ou_atualizar_auditoria(registro)
                    st.success("🗄️ Salvo! Se foi REPROVADA, acesse a aba 'Anexar Evidências' para subir o print.")
                    st.rerun()



        # ── Histórico somente na aba Realizar Auditoria ───────────────────────
        id_cliente_hist = st.session_state.get("id_cliente_selecionado", None)
        exibir_historico_auditoria_agrupado(df_salvos, id_cliente_hist)

        # ── PDF das reprovadas do filtro atual, gerado apenas sob demanda ─────
        exibir_pdf_reprovadas_auditoria(df_salvos, id_cliente_hist)

    # =========================================================================
    # ABA 2: ANEXAR / SUBSTITUIR EVIDÊNCIAS
    # =========================================================================
    with tab_evidencias:
        st.subheader("📸 Fila de Evidências (Câmeras Reprovadas)")
        st.caption("Agora é possível anexar, substituir ou remover evidência mesmo depois do primeiro cadastro.")

        if not df_salvos.empty:
            df_reprovadas = garantir_colunas_dataframe(carregar_reprovadas(df_salvos), ["ID_Whitelabel", "Franqueado", "ID_da_Camera", "Nome_da_Camera", "Caminho_Evidencia", "Caminho_Thumbnail", "Resultado_Geral", "Observacoes"])

            if df_reprovadas.empty:
                st.success("🟢 Nenhuma câmera está marcada como REPROVADA no momento. Ótimo trabalho!")
            else:
                id_cliente_selecionado = st.session_state.get("id_cliente_selecionado", None)
                df_evid = df_reprovadas.copy()

                if id_cliente_selecionado:
                    df_evid = df_evid[df_evid["ID_Whitelabel"].astype(str).str.strip() == str(id_cliente_selecionado).strip()]
                    if df_evid.empty:
                        st.warning("Esse cliente selecionado não possui câmeras reprovadas. Mostrando todas as reprovadas.")
                        df_evid = df_reprovadas.copy()

                col_busca1, col_busca2, col_busca3 = st.columns([2, 1, 1])
                with col_busca1:
                    busca_evid = st.text_input("Buscar câmera reprovada:", key="busca_evidencias", placeholder="ID, nome da câmera ou franqueado")
                with col_busca2:
                    itens_por_pagina = st.selectbox("Itens por página", [10, 25, 50, 100], index=1, key="evid_itens_pagina")
                with col_busca3:
                    somente_sem_foto = st.checkbox("Somente sem foto", key="evid_somente_sem_foto")

                if busca_evid:
                    termo = busca_evid.lower().strip()
                    df_evid = df_evid[
                        df_evid["ID_da_Camera"].astype(str).str.lower().str.contains(termo, na=False) |
                        df_evid["Nome_da_Camera"].astype(str).str.lower().str.contains(termo, na=False) |
                        df_evid["Franqueado"].astype(str).str.lower().str.contains(termo, na=False)
                    ]

                if somente_sem_foto:
                    df_evid = df_evid[df_evid["ID_da_Camera"].astype(str).apply(lambda x: contar_evidencias(x) == 0)]

                total_evid = len(df_evid)
                total_paginas = max(1, (total_evid + itens_por_pagina - 1) // itens_por_pagina)
                pagina_atual = st.number_input("Página", min_value=1, max_value=total_paginas, value=1, step=1, key="evid_pagina")
                inicio = (pagina_atual - 1) * itens_por_pagina
                fim = inicio + itens_por_pagina
                df_evid_pagina = df_evid.iloc[inicio:fim].copy()

                st.caption(f"Mostrando {len(df_evid_pagina)} de {total_evid} câmeras reprovadas.")

                opcoes_reprovadas = []
                for _, r in df_evid_pagina.iterrows():
                    total_fotos = contar_evidencias(r["ID_da_Camera"])
                    if total_fotos == 0:
                        img_verificar = r.get("Caminho_Evidencia", "")
                        thumb_verificar = r.get("Caminho_Thumbnail", "")
                        migrar_evidencia_antiga_para_galeria(r["ID_da_Camera"], img_verificar, thumb_verificar)
                        total_fotos = contar_evidencias(r["ID_da_Camera"])
                    tem_foto = f"📸 {total_fotos} foto(s)" if total_fotos > 0 else "❌ Sem foto"
                    opcoes_reprovadas.append(f"{r['ID_da_Camera']} - {r['Nome_da_Camera']} ({r['Franqueado']}) | {tem_foto}")

                camera_reprovada_sel = st.selectbox(
                    "Escolha a câmera reprovada para gerenciar a evidência:",
                    options=[""] + opcoes_reprovadas,
                    key="sel_camera_evid_v3"
                )

                if camera_reprovada_sel:
                    id_reprovada = camera_reprovada_sel.split(" - ")[0]
                    dados_camera_rep = df_reprovadas[df_reprovadas["ID_da_Camera"].astype(str) == str(id_reprovada)].iloc[0]

                    st.info(f"📋 **Câmera:** {dados_camera_rep['Nome_da_Camera']} | **Franqueado:** {dados_camera_rep['Franqueado']}")

                    img_atual = dados_camera_rep.get("Caminho_Evidencia", "")
                    caminho_foto_atual = str(img_atual).strip() if pd.notna(img_atual) else ""
                    thumb_atual = str(dados_camera_rep.get("Caminho_Thumbnail", "")).strip() if pd.notna(dados_camera_rep.get("Caminho_Thumbnail", "")) else ""

                    # Migra automaticamente a imagem antiga para a nova galeria de evidências.
                    migrar_evidencia_antiga_para_galeria(id_reprovada, caminho_foto_atual, thumb_atual)
                    evidencias_camera = listar_evidencias(id_reprovada)
                    evidencias_validas = [
                        ev for ev in evidencias_camera
                        if str(ev.get("Caminho_Evidencia", "")).strip() and caminho_existe(str(ev.get("Caminho_Evidencia", "")).strip())
                    ]

                    @st.dialog("🔍 Evidência em tamanho maior", width="large")
                    def abrir_popup_evidencia_fila(caminho_img, idx_evidencia, id_camera, nome_camera, cliente, cidade, uf, data_upload, obs_foto="", obs_auditoria=""):
                        obs_foto = str(obs_foto or "").strip()
                        obs_auditoria = str(obs_auditoria or "").strip()

                        st.info(
                            f"**Evidência:** {idx_evidencia}  \n"
                            f"**ID da Câmera:** {id_camera}  \n"
                            f"**Nome da Câmera:** {nome_camera}  \n"
                            f"**Cliente:** {cliente}  \n"
                            f"**Cidade/UF:** {cidade}/{uf}  \n"
                            f"**Data do upload:** {data_upload}  \n\n"
                            f"**Observação Auditoria:** {obs_auditoria if obs_auditoria else 'Sem observação'}"
                        )

                        if caminho_img and caminho_existe(caminho_img):
                            st.image(
                                caminho_img,
                                use_container_width=True,
                                caption=f"Evidência {idx_evidencia} - Câmera {id_camera}"
                            )
                        else:
                            st.warning("A imagem selecionada não foi encontrada no disco.")

                        st.divider()
                        st.markdown("### 📝 Observação da Evidência")
                        if obs_foto:
                            st.info(obs_foto)
                        else:
                            st.info("Sem observação cadastrada para esta evidência.")

                        # Observação da auditoria exibida junto ao bloco azul acima.

                    if evidencias_validas:
                        st.markdown(f"### 🖼️ Evidências anexadas ({len(evidencias_validas)})")
                        cols_galeria = st.columns(3)

                        for idx_ev, ev in enumerate(evidencias_validas, start=1):
                            caminho_ev = str(ev.get("Caminho_Evidencia", "")).strip()
                            thumb_ev = str(ev.get("Caminho_Thumbnail", "")).strip()
                            data_ev = ev.get("Data_Upload", "") or "Sem data"
                            obs_ev = ev.get("Observacao", "") or ""
                            id_ev = ev.get("ID")

                            with cols_galeria[(idx_ev - 1) % 3]:
                                if thumb_ev and caminho_existe(thumb_ev):
                                    st.image(thumb_ev, caption=f"Evidência {idx_ev} - {data_ev}", width=260)
                                else:
                                    imagem_preview = carregar_imagem_otimizada(caminho_ev, 420)
                                    if imagem_preview:
                                        st.image(imagem_preview, caption=f"Evidência {idx_ev} - {data_ev}", width=260)

                                col_ev1, col_ev2 = st.columns(2)
                                with col_ev1:
                                    if st.button("🔍 Ver", key=f"btn_ver_ev_{id_ev}"):
                                        abrir_popup_evidencia_fila(
                                            caminho_ev,
                                            idx_ev,
                                            id_reprovada,
                                            dados_camera_rep.get("Nome_da_Camera", ""),
                                            dados_camera_rep.get("Franqueado", ""),
                                            dados_camera_rep.get("Cidade", ""),
                                            dados_camera_rep.get("UF", ""),
                                            data_ev,
                                            obs_ev,
                                            dados_camera_rep.get("Observacoes", "")
                                        )

                                with col_ev2:
                                    if st.button("🗑️ Excluir", key=f"btn_del_ev_{id_ev}"):
                                        dados_removidos = remover_evidencia_unica(id_ev)
                                        if dados_removidos:
                                            excluir_arquivo_se_existir(dados_removidos.get("Caminho_Evidencia", ""))
                                            excluir_arquivo_se_existir(dados_removidos.get("Caminho_Thumbnail", ""))
                                        st.cache_data.clear()
                                        st.success("Evidência removida.")
                                        st.rerun()
                    else:
                        st.warning("Essa câmera ainda não possui evidência anexada.")

                    st.divider()
                    st.markdown("### 📤 Adicionar novas evidências")
                    st.caption("Este bloco **não substitui** imagens existentes. Cada arquivo enviado vira uma nova imagem na galeria da câmera.")

                    if f"upload_reset_{id_reprovada}" not in st.session_state:
                        st.session_state[f"upload_reset_{id_reprovada}"] = 0

                    with st.form(key=f"form_upload_evid_{id_reprovada}_{st.session_state[f'upload_reset_{id_reprovada}']}", clear_on_submit=True):
                        arquivos_upload = st.file_uploader(
                            "Adicionar um ou mais prints de evidência:",
                            type=["png", "jpg", "jpeg"],
                            accept_multiple_files=True,
                            key=f"f_up_v33_{id_reprovada}_{st.session_state[f'upload_reset_{id_reprovada}']}"
                        )
                        observacao_evidencia = st.text_input(
                            "Observação para as novas imagens (opcional):",
                            key=f"obs_ev_{id_reprovada}_{st.session_state[f'upload_reset_{id_reprovada}']}"
                        )

                        st.caption("Ao salvar, as imagens são acrescentadas à galeria, comprimidas automaticamente e recebem thumbnails persistidas.")

                        salvar_upload = st.form_submit_button(
                            "➕ Adicionar à galeria",
                            type="primary"
                        )

                    if salvar_upload:
                        if not arquivos_upload:
                            st.warning("Selecione pelo menos uma imagem antes de salvar.")
                        else:
                            qtd_adicionada = 0
                            with st.spinner("Comprimindo imagens, criando thumbnails e adicionando à galeria..."):
                                for arquivo_upload in arquivos_upload:
                                    caminho_final_foto, caminho_thumb = salvar_evidencia_otimizada(arquivo_upload, id_reprovada)
                                    adicionar_evidencia(id_reprovada, caminho_final_foto, caminho_thumb, observacao_evidencia)
                                    qtd_adicionada += 1

                            st.cache_data.clear()
                            st.session_state[f"upload_reset_{id_reprovada}"] += 1
                            st.success(f"✅ {qtd_adicionada} nova(s) evidência(s) adicionada(s) à galeria!")
                            st.rerun()
        else:
            st.info("ℹ️ Nenhuma câmera auditada no banco de dados ainda.")


    # =========================================================================
    # ABA 3: REPROVADAS AGRUPADAS POR CLIENTE
    # =========================================================================
    with tab_agrupamento:
        st.subheader("📊 Câmeras Reprovadas Agrupadas por Cliente")
        st.caption("Esta tela substitui o histórico geral visual nessa área e mostra somente as câmeras reprovadas, agrupadas por cliente.")

        df_reprovadas_agrupado = preparar_df_reprovadas_agrupado(df_salvos)

        if df_reprovadas_agrupado.empty:
            st.success("🟢 Nenhuma câmera reprovada encontrada no momento.")
        else:
            resumo_clientes = []
            for (id_wl, franq), grp in df_reprovadas_agrupado.groupby(["ID_Whitelabel", "Franqueado"], dropna=False):
                total_evid = sum(contar_evidencias_camera(row.to_dict()) for _, row in grp.iterrows())
                resumo_clientes.append({
                    "ID_Whitelabel": id_wl,
                    "Franqueado": franq,
                    "Cidade/UF": f"{grp.iloc[0].get('Cidade', '')}/{grp.iloc[0].get('UF', '')}",
                    "Câmeras Reprovadas": len(grp),
                    "Evidências": total_evid,
                })

            df_resumo_clientes = pd.DataFrame(resumo_clientes).sort_values(["Franqueado", "ID_Whitelabel"])

            col_ag1, col_ag2, col_ag3 = st.columns(3)
            col_ag1.metric("Clientes com Reprovação", len(df_resumo_clientes))
            col_ag2.metric("Câmeras Reprovadas", len(df_reprovadas_agrupado))
            col_ag3.metric("Evidências Anexadas", int(df_resumo_clientes["Evidências"].sum()))

            st.markdown("### Resumo por Cliente")
            st.dataframe(df_resumo_clientes, use_container_width=True, hide_index=True)

            st.markdown("### Detalhamento")
            clientes_para_filtro = ["Todos"] + [f"{r['ID_Whitelabel']} - {r['Franqueado']}" for _, r in df_resumo_clientes.iterrows()]
            filtro_cliente_rep = st.selectbox("Filtrar cliente", clientes_para_filtro, key="filtro_cliente_reprovadas_agrupadas")

            df_detalhe = df_reprovadas_agrupado.copy()
            if filtro_cliente_rep != "Todos":
                id_filtrado = filtro_cliente_rep.split(" - ")[0].strip()
                df_detalhe = df_detalhe[df_detalhe["ID_Whitelabel"].astype(str) == id_filtrado]

            for (id_wl, franq), grp in df_detalhe.groupby(["ID_Whitelabel", "Franqueado"], dropna=False):
                with st.expander(f"{id_wl} - {franq} | {len(grp)} reprovada(s)", expanded=(filtro_cliente_rep != "Todos")):
                    df_exibir = grp.copy()
                    df_exibir["Qtd Evidências"] = df_exibir.apply(lambda row: contar_evidencias_camera(row.to_dict()), axis=1)
                    colunas_exibir = [
                        "ID_da_Camera", "Nome_da_Camera", "Cidade", "UF", "Status_da_Camera",
                        "Marca d'Água Travada", "Câmera está com um bom foco",
                        "Câmera está bem posicionada", "LPR lendo de forma efetiva",
                        "Qtd Evidências", "Observacoes"
                    ]
                    colunas_exibir = [c for c in colunas_exibir if c in df_exibir.columns]
                    st.dataframe(df_exibir[colunas_exibir], use_container_width=True, hide_index=True)

                    st.markdown("**Galeria resumida**")
                    st.caption("Clique em **Abrir Imagem** abaixo da miniatura para visualizar a evidência em tamanho maior, em uma janela flutuante.")

                    @st.dialog("🔍 Evidência em tamanho maior", width="large")
                    def abrir_modal_evidencia(caminho_img, id_camera, nome_camera, cliente, data_upload, obs_foto="", obs_auditoria=""):
                        obs_foto = str(obs_foto or "").strip()
                        obs_auditoria = str(obs_auditoria or "").strip()

                        st.info(
                            f"**Câmera:** {id_camera} - {nome_camera}  \n"
                            f"**Cliente:** {cliente}  \n"
                            f"**Data do upload:** {data_upload}"
                        )

                        if caminho_img and caminho_existe(caminho_img):
                            st.image(
                                caminho_img,
                                use_container_width=True,
                                caption="Imagem original da evidência"
                            )
                        else:
                            st.warning("A imagem selecionada não foi encontrada no disco.")

                        st.divider()
                        st.markdown("### 📝 Observação da Evidência")
                        if obs_foto:
                            st.info(obs_foto)
                        else:
                            st.info("Sem observação cadastrada para esta evidência.")

                        # Observação da auditoria exibida junto ao bloco azul acima.

                    for _, row in grp.iterrows():
                        evidencias = listar_evidencias(row.get("ID_da_Camera", ""))
                        itens_galeria = []

                        for ev in evidencias:
                            thumb = str(ev.get("Caminho_Thumbnail", "")).strip()
                            img = str(ev.get("Caminho_Evidencia", "")).strip()

                            if img and caminho_existe(img):
                                itens_galeria.append({
                                    "thumb": thumb if thumb and caminho_existe(thumb) else img,
                                    "original": img,
                                    "id_evidencia": ev.get("ID", ""),
                                    "data_upload": ev.get("Data_Upload", ""),
                                    "obs": ev.get("Observacao", ""),
                                })

                        if itens_galeria:
                            st.caption(f"📷 {row.get('ID_da_Camera', '')} - {row.get('Nome_da_Camera', '')}")
                            cols_thumb = st.columns(min(4, len(itens_galeria)))

                            for idx_thumb, item_galeria in enumerate(itens_galeria):
                                col_thumb = cols_thumb[idx_thumb % len(cols_thumb)]

                                with col_thumb:
                                    st.image(item_galeria["thumb"], width=150)

                                    if st.button(
                                        "🔍 Abrir Imagem",
                                        key=f"abrir_popup_galeria_{row.get('ID_da_Camera', '')}_{idx_thumb}_{item_galeria.get('id_evidencia', '')}"
                                    ):
                                        abrir_modal_evidencia(
                                            item_galeria["original"],
                                            row.get("ID_da_Camera", ""),
                                            row.get("Nome_da_Camera", ""),
                                            row.get("Franqueado", ""),
                                            item_galeria.get("data_upload", ""),
                                            item_galeria.get("obs", ""),
                                            row.get("Observacoes", "")
                                        )

    # ── Barra Lateral Dinâmica (Métricas e Exportação 100% Reativas) ──
    with st.sidebar:
        st.header("📊 Painel de Métricas")
        if st.button("🔄 Recarregar arquivos de origem", key="btn_recarregar_origem"):
            st.cache_data.clear()
            st.rerun()

        if not df_salvos.empty and id_cliente_selecionado:
            id_filtro_sidebar = str(id_cliente_selecionado).strip()
            df_sidebar = df_salvos[df_salvos["ID_Whitelabel"] == id_filtro_sidebar]
            st.caption(f"Visualizando métricas do cliente: **{id_filtro_sidebar}**")
        else:
            df_sidebar = df_salvos
            st.caption("Visualizando métricas de **Todos os Clientes**")

        total_registros = len(df_sidebar) if not df_sidebar.empty else 0
        total_aprovadas = (df_sidebar["Resultado_Geral"] == "APROVADA").sum() if total_registros > 0 else 0
        total_reprovadas = (df_sidebar["Resultado_Geral"] == "REPROVADA").sum() if total_registros > 0 else 0

        st.metric("Total no Banco SQL", total_registros)
        st.metric("🟢 Câmeras Aprovadas", total_aprovadas)
        st.metric("❌ Câmeras Reprovadas", total_reprovadas)

        if total_registros > 0:
            st.divider()
            st.subheader("Exportação das Reprovadas")
            st.caption("Excel e PDF serão gerados somente com câmeras REPROVADAS, agrupadas por cliente, e apenas quando você clicar no botão.")

            sufixo_arquivo = f"cliente_{id_cliente_selecionado}" if id_cliente_selecionado else "geral"
            chave_relatorio = f"relatorio_{sufixo_arquivo}_{total_registros}_{total_reprovadas}"

            if st.session_state.get("chave_relatorio_excel") != chave_relatorio:
                st.session_state.pop("excel_data_pronto", None)
                st.session_state.pop("excel_nome_arquivo", None)
                st.session_state["chave_relatorio_excel"] = chave_relatorio

            if st.button("📄 Preparar Excel das Reprovadas", type="primary"):
                with st.spinner("Montando relatório Excel..."):
                    excel_data = gerar_excel_reprovadas_agrupadas(df_sidebar)
                    nome_arquivo = f"reprovadas_por_cliente_{sufixo_arquivo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    st.session_state["excel_data_pronto"] = excel_data
                    st.session_state["excel_nome_arquivo"] = nome_arquivo

            if "excel_data_pronto" in st.session_state:
                st.download_button(
                    label="⬇️ Baixar Excel das Reprovadas",
                    data=st.session_state["excel_data_pronto"],
                    file_name=st.session_state.get("excel_nome_arquivo", "auditoria.xlsx"),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            st.divider()
            st.subheader("Relatório PDF das Reprovadas")
            st.caption("O PDF também só é montado sob demanda, para evitar travamento.")
            chave_pdf = f"pdf_{sufixo_arquivo}_{total_registros}_{total_reprovadas}"
            if st.session_state.get("chave_relatorio_pdf") != chave_pdf:
                st.session_state.pop("pdf_data_pronto", None)
                st.session_state.pop("pdf_nome_arquivo", None)
                st.session_state["chave_relatorio_pdf"] = chave_pdf

            if st.button("🧾 Preparar PDF das Reprovadas", type="secondary"):
                try:
                    with st.spinner("Montando relatório PDF..."):
                        pdf_data = gerar_pdf_reprovadas_agrupadas(df_sidebar)
                        nome_pdf = f"reprovadas_por_cliente_{sufixo_arquivo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                        st.session_state["pdf_data_pronto"] = pdf_data
                        st.session_state["pdf_nome_arquivo"] = nome_pdf
                except Exception as e:
                    st.error(f"Erro ao gerar PDF: {e}")

            if "pdf_data_pronto" in st.session_state:
                st.download_button(
                    label="⬇️ Baixar PDF das Reprovadas",
                    data=st.session_state["pdf_data_pronto"],
                    file_name=st.session_state.get("pdf_nome_arquivo", "auditoria.pdf"),
                    mime="application/pdf"
                )

if __name__ == "__main__":
    main()
